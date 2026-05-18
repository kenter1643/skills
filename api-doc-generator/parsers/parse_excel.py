# -*- coding: utf-8 -*-
"""
Excel 解析器
仅解析"字段说明"sheet，提取全部接口数据
"""

import openpyxl
import sys
import os

_skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _skill_root not in sys.path:
    sys.path.insert(0, _skill_root)

import utils


def _find_col_indices(ws):
    """
    动态扫描字段说明 sheet 的表头行，返回关键列的索引（0-based）。
    扫描前 10 行，找包含关键词的行作为表头。
    """
    keyword_map = {
        '字段名称': 'field_name',
        '是否必填': 'is_required',
        '是否枚举': 'is_enum',
        '枚举内容': 'enum_values',
        '是否列表字段': 'is_list_field',
        '列表顺序': 'list_order',
        '查询方式': 'query_type',
        '字段类型': 'field_type',
    }
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        indices = {}
        for col_idx, cell_val in enumerate(row):
            if not cell_val:
                continue
            cell_str = str(cell_val).strip()
            for keyword, key in keyword_map.items():
                if keyword in cell_str:
                    indices[key] = col_idx
        if 'field_name' in indices:
            return indices
    return None


def parse_excel(excel_path):
    """
    解析 Excel 文件，仅从"字段说明"sheet获取全部数据

    Args:
        excel_path: Excel 文件路径

    Returns:
        dict: {
            'fuzzy_search': [字段中文名, ...],
            'advanced_search': [字段中文名, ...],
            'list_fields': [(字段中文名, 顺序), ...],  # 带顺序的列表字段
            'field_info': {字段中文名: {is_required, is_enum, enum_values}},
            'has_attachment': bool,
            'field_order': [字段中文名, ...],  # 按Excel中的顺序
        }
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 查找"字段说明"sheet
    field_sheet_name = None
    for sheet_name in wb.sheetnames:
        if '字段说明' in sheet_name:
            field_sheet_name = sheet_name
            break

    if not field_sheet_name:
        raise ValueError(f"未找到字段说明 sheet，请确认 Excel 中包含 '字段说明' 的 sheet")

    ws = wb[field_sheet_name]
    col_indices = _find_col_indices(ws)
    if col_indices is None:
        raise ValueError("字段说明 sheet 中未找到表头，请确认包含'字段名称'列")

    fn_col = col_indices.get('field_name')
    req_col = col_indices.get('is_required')
    enum_col = col_indices.get('is_enum')
    eval_col = col_indices.get('enum_values')
    list_col = col_indices.get('is_list_field')
    order_col = col_indices.get('list_order')
    query_col = col_indices.get('query_type')
    type_col = col_indices.get('field_type')

    # 定位表头行号
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if fn_col is not None and fn_col < len(row) and row[fn_col] and '字段名称' in str(row[fn_col]):
            header_row = row_idx
            break
    data_start = (header_row + 1) if header_row else 2

    fuzzy_search = []
    advanced_search = []
    list_fields = []  # (中文名, 顺序)
    field_info = {}
    field_order = []
    has_attachment = False
    fuzzy_seen = set()
    advanced_seen = set()
    list_seen = set()
    in_detail_section = False

    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start, values_only=False), start=data_start):
        # 检查第一列（A列）是否以"明细"开头，用于识别明细区域标题行
        first_cell = row[0].value if len(row) > 0 else None
        if first_cell and str(first_cell).strip().startswith('明细'):
            in_detail_section = True
            continue  # 跳过明细标题行本身

        # 空行结束明细区域
        if not any(cell.value for cell in row):
            in_detail_section = False
            continue

        if fn_col is None or fn_col >= len(row):
            continue

        raw_name = row[fn_col].value
        if not raw_name:
            continue

        field_name = utils.canonicalize_field_name(raw_name)

        # 是否必填
        is_required = False
        if req_col is not None and req_col < len(row):
            val = row[req_col].value
            is_required = str(val).strip() == '是' if val else False

        # 是否枚举
        is_enum = False
        if enum_col is not None and enum_col < len(row):
            val = row[enum_col].value
            is_enum = str(val).strip() == '是' if val else False

        # 枚举内容
        enum_values = ''
        if eval_col is not None and eval_col < len(row):
            enum_values = row[eval_col].value or ''

        # 是否列表字段
        is_list_field = False
        if list_col is not None and list_col < len(row):
            val = row[list_col].value
            is_list_field = str(val).strip() == '是' if val else False

        # 列表顺序
        list_order = None
        if order_col is not None and order_col < len(row):
            val = row[order_col].value
            if val is not None:
                try:
                    list_order = int(val)
                except (ValueError, TypeError):
                    list_order = None

        # 查询方式
        query_type = None
        if query_col is not None and query_col < len(row):
            val = row[query_col].value
            if val:
                query_type = str(val).strip()

        # 字段类型（兜底用）
        excel_field_type = None
        if type_col is not None and type_col < len(row):
            val = row[type_col].value
            if val:
                excel_field_type = str(val).strip()

        # 检查是否包含附件
        if '附件' in str(raw_name) or '文件' in str(raw_name):
            has_attachment = True

        # 处理查询方式
        if query_type == '模糊':
            if field_name not in fuzzy_seen:
                fuzzy_seen.add(field_name)
                fuzzy_search.append(field_name)
            if field_name not in advanced_seen:
                advanced_seen.add(field_name)
                advanced_search.append(field_name)
        elif query_type == '高级':
            if field_name not in advanced_seen:
                advanced_seen.add(field_name)
                advanced_search.append(field_name)

        # 处理列表字段
        if is_list_field:
            if field_name not in list_seen:
                list_seen.add(field_name)
                list_fields.append((field_name, list_order))

        # 保存字段信息
        field_info[field_name] = {
            'is_required': is_required,
            'is_enum': is_enum,
            'enum_values': str(enum_values),
            'excel_field_type': excel_field_type,
            'is_detail_section': in_detail_section,
        }

        # 保留顺序
        if field_name not in field_order:
            field_order.append(field_name)

    # 列表字段排序：有顺序的在前（按数字升序），无顺序的在后
    list_fields_with_order = [(name, order) for name, order in list_fields if order is not None]
    list_fields_with_order.sort(key=lambda x: x[1])
    list_fields_no_order = [(name, None) for name, order in list_fields if order is None]
    sorted_list_fields = list_fields_with_order + list_fields_no_order

    return {
        'fuzzy_search': fuzzy_search,
        'advanced_search': advanced_search,
        'list_fields': sorted_list_fields,  # [(中文名, 顺序), ...]
        'field_info': field_info,
        'has_attachment': has_attachment,
        'field_order': field_order,
    }


if __name__ == '__main__':
    path = r"D:\ZY\code\jx\construct-star-server-v1.10.9sp9\product\项目收入科目登记.xlsx"
    data = parse_excel(path)
    print(f"模糊查询: {data['fuzzy_search']}")
    print(f"高级查询: {data['advanced_search']}")
    print(f"列表字段(带顺序): {data['list_fields']}")
    print(f"字段说明数量: {len(data['field_info'])}")
    print(f"字段顺序: {data['field_order'][:10]}...")
