# -*- coding: utf-8 -*-
import openpyxl, sys, os, json, re

excel_path = os.path.join(os.path.dirname(__file__), "..", "source", "建星业财V2.0.1（2）-成本管理业务-收入线业务模块概要需求分析.xlsx")
wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb['03-合同内(施工图)预算变更-XXX-字段说明']

print(f"max_row={ws.max_row}, max_col={ws.max_column}")
print("\nAll rows:")
for row_idx in range(1, ws.max_row+1):
    row_vals = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(row=row_idx, column=c).value
        row_vals.append(str(v)[:40] if v is not None else "")
    print(f"  Row {row_idx:2d}: {row_vals}")
