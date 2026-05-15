# -*- coding: utf-8 -*-
import openpyxl
import sys

path = sys.argv[1] if len(sys.argv) > 1 else None
if not path:
    print("Usage: python inspect_excel.py <excel_path>")
    sys.exit(1)

wb = openpyxl.load_workbook(path, data_only=True)
print("=== Sheets ===")
for name in wb.sheetnames:
    print(f"  - {name}")

# 找字段说明 sheet
field_sheet = None
for name in wb.sheetnames:
    if '字段说明' in name:
        field_sheet = name
        break

if field_sheet:
    print(f"\n=== Sheet: {field_sheet} (前50行) ===")
    ws = wb[field_sheet]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), 1):
        non_empty = [str(c) for c in row if c is not None]
        if non_empty:
            print(f"  Row {i}: {non_empty}")
