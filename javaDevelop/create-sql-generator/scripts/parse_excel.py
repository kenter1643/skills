# -*- coding: utf-8 -*-
"""
解析 Excel 需求文档，提取"字段说明" sheet 的所有字段行。

用法:
    python parse_excel.py <excel_path>

输出: JSON 到 stdout，field_name 为 field_label 的 snake_case 英文翻译。
"""

import sys
import json
import re
import os
import openpyxl


# ── 类型映射 ─────────────────────────────────────────────────────────────────

TYPE_MAP = {
    # 整型
    "int": "int",
    "整型": "int",
    "整数": "int",
    "integer": "int",
    "tinyint": "tinyint",
    "bigint": "bigint",
    # 小数
    "decimal": "decimal(24,8)",
    "金额": "decimal(24,8)",
    "数值": "decimal(24,8)",
    "float": "decimal(24,8)",
    "double": "decimal(24,8)",
    # 字符串
    "varchar": "varchar",
    "字符": "varchar",
    "string": "varchar",
    "text": "text",
    "longtext": "longtext",
    # 时间
    "datetime": "datetime",
    "date": "datetime",
    "时间": "datetime",
    "日期": "datetime",
    # 布尔/状态
    "boolean": "tinyint",
    "bool": "tinyint",
    "是否": "tinyint",
}

DEFAULT_VARCHAR_LENGTH = 200


def resolve_type(raw_type: str, field_label: str = "") -> str:
    """将 Excel 中的类型描述转换为 MySQL 类型。"""
    if not raw_type:
        raw_type = ""
    t = raw_type.strip().lower()

    # 已经是标准 MySQL 类型（含长度），直接返回
    if re.match(r"^(varchar|char)\s*\(\d+\)", t):
        return raw_type.strip()
    if re.match(r"^decimal\s*\(\d+,\d+\)", t):
        return raw_type.strip()

    for key, mapped in TYPE_MAP.items():
        if key in t:
            if mapped == "varchar":
                # 尝试从括号提取长度
                m = re.search(r"\((\d+)\)", raw_type)
                length = m.group(1) if m else str(DEFAULT_VARCHAR_LENGTH)
                return f"varchar({length})"
            return mapped

    # 根据字段名推断
    label = field_label.lower()
    if any(k in label for k in ["金额", "价格", "单价", "税率", "数量"]):
        return "decimal(24,8)"
    if any(k in label for k in ["时间", "日期", "年月"]):
        return "datetime"
    if any(k in label for k in ["是否", "状态", "类型", "属性"]) and len(field_label) <= 6:
        return "tinyint"
    if any(k in label for k in ["id", "编号", "编码"]) and "名" not in label:
        return "varchar(50)"
    if any(k in label for k in ["名称", "姓名"]):
        return "varchar(200)"
    if any(k in label for k in ["说明", "备注", "描述"]):
        return "varchar(2000)"

    # 默认 varchar(200)
    return f"varchar({DEFAULT_VARCHAR_LENGTH})"


def to_snake_case(name: str) -> str:
    """
    将字段名转换为 snake_case 小写下划线格式。
    规则：
    - 纯大写缩写（如 FXMBH、FSSZZ）→ 直接转小写：fxmbh、fsszz
    - 驼峰（如 FPinPai、FYear_FMonth）→ 插入下划线后转小写：f_pin_pai、f_year_f_month
    - 已是 snake_case（如 project_id）→ 保持不变，只转小写
    - 前缀 F 开头的浪潮字段（如 FRKSL、FRate）→ 转小写：frksl、f_rate
    """
    # 已含下划线，直接转小写
    if "_" in name:
        return name.lower()
    # 全大写（如 FXMBH）→ 直接转小写
    if name.isupper():
        return name.lower()
    # 驼峰转下划线：在大写字母前插入下划线（连续大写视为缩写，不拆分）
    # 例：FPinPai → F_Pin_Pai → f_pin_pai
    #     FRate   → F_Rate    → f_rate
    s = re.sub(r"([A-Z]+)([A-Z][a-z])", r"\1_\2", name)  # 缩写+单词边界
    s = re.sub(r"([a-z\d])([A-Z])", r"\1_\2", s)          # 小写/数字+大写
    return s.lower()


# ── sheet 解析 ────────────────────────────────────────────────────────────────

def find_field_sheets(wb):
    """返回所有包含"字段说明"的 sheet 名列表。"""
    return [name for name in wb.sheetnames if "字段说明" in name]


def find_header_row(ws):
    """
    找表头行：扫描前 15 行，任意列包含"字段名"、"英文"、"编号"且同行有"字段类型"或"分组"。
    """
    HEADER_KEYWORDS = {"字段名", "字段列名", "英文字段", "字段名称"}
    CONFIRM_KEYWORDS = {"字段类型", "分组", "是否必填", "枚举"}
    for row_idx in range(1, 16):
        row_vals = [str(ws.cell(row=row_idx, column=c).value or "").strip()
                    for c in range(1, min(ws.max_column + 1, 20))]
        has_header = any(any(k in v for k in HEADER_KEYWORDS) for v in row_vals)
        has_confirm = any(any(k in v for k in CONFIRM_KEYWORDS) for v in row_vals)
        if has_header and has_confirm:
            return row_idx
    return None


def parse_col_indices(ws, header_row: int) -> dict:
    """
    扫描表头行，返回关键列的 1-based 列号。
    支持多种表头命名方式，包括"字段名称"+"字段列名（筑云）"格式。
    """
    # 优先级顺序：列表中越靠前的关键词越优先，一旦匹配就停止
    # 注意：field_name 和 field_label 的关键词不能互相包含，避免同一列被两个 key 争抢
    mapping = {
        "field_name":  ["字段英文名", "字段列名", "英文字段名", "fieldname", "field_name"],
        "field_label": ["字段中文名", "中文名称", "字段名称", "中文名"],
        "field_type":  ["字段类型", "类型", "type"],
        "group":       ["功能模块信息", "分组信息", "分组", "group"],
        "length":      ["长度", "length"],
        "required":    ["是否必填", "必填", "required"],
        "enum_values": ["枚举属性值", "枚举内容", "枚举值", "枚举", "可选值", "enum"],
        "remark":      ["字段说明", "备注", "说明", "remark"],
    }

    result = {}
    # 先按关键词优先级从高到低扫描所有列，找到最优匹配
    # 为了避免一个列头同时被多个 key 匹配，每列只分配给优先级最高的 key
    col_assignments: dict[int, tuple[str, int]] = {}  # col -> (key, priority)
    row = ws[header_row]
    for cell in row:
        if not cell.value:
            continue
        val = str(cell.value).strip().lower()
        for key, keywords in mapping.items():
            for pri, kw in enumerate(keywords):
                if kw.lower() in val:
                    prev = col_assignments.get(cell.column)
                    if prev is None or pri < prev[1]:
                        col_assignments[cell.column] = (key, pri)
                    break

    # 每个 key 取优先级最高（pri 最小）的列
    key_best: dict[str, tuple[int, int]] = {}  # key -> (col, pri)
    for col_no, (key, pri) in col_assignments.items():
        prev = key_best.get(key)
        if prev is None or pri < prev[1]:
            key_best[key] = (col_no, pri)

    result = {key: col_no for key, (col_no, _) in key_best.items()}
    return result


def detect_table_section(cell_value: str) -> str | None:
    """
    判断某行是否是分区标题（主表 / 明细表 / 明细表#xxx）。
    返回标准化的 table_type 字符串，或 None。
    """
    if not cell_value:
        return None
    v = str(cell_value).strip()

    # 匹配形如 "明细表#资产信息" 或 "明细表（资产信息）"
    m = re.match(r"(明细表)[#（(](.+?)[）)]?$", v)
    if m:
        return f"明细表#{m.group(2)}"

    # 匹配形如 "附表1-合同项目明细"、"附表2-资产信息" → 视为 明细表#XXX
    m2 = re.match(r"附表\d+[-—]\s*(.+)$", v)
    if m2:
        return f"明细表#{m2.group(1).strip()}"

    if v in ("主表", "head", "HEAD"):
        return "主表"
    if v in ("明细表", "detail", "DETAIL"):
        return "明细表"
    if v in ("info", "INFO", "info表", "INFO表"):
        return "info表"

    # 宽松匹配：整行只有这几个词
    if re.fullmatch(r"[主明细Ii][表nfoNFO表细]*", v):
        return v

    return None


def parse_field_sheet(ws, sheet_name: str) -> list:
    """
    解析单个字段说明 sheet，返回字段列表。
    支持两种分区方式：
    1. 独立分区标题行（整行只有一个值如"主表"/"明细表"/"明细表#XXX"）
    2. 分组列（group 列）：表头分组名决定归属，"表头"→主表，其余→明细表
    """
    header_row = find_header_row(ws)
    if header_row is None:
        return []

    col = parse_col_indices(ws, header_row)
    if "field_name" not in col:
        return []

    fields = []
    current_table_type = "明细表"  # 默认归明细表
    last_group = ""               # 记录上一个非空分组名，用于跨行继承

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # 取第一列值，判断是否是独立分区标题行
        first_cell = ws.cell(row=row_idx, column=1).value
        section = detect_table_section(first_cell)
        if section:
            current_table_type = section
            last_group = ""
            continue

        # 取字段列名（英文），若为空则回退到字段名称列（中文，后续翻译）
        fn_cell = ws.cell(row=row_idx, column=col["field_name"]).value
        fl_cell = ws.cell(row=row_idx, column=col["field_label"]).value if "field_label" in col else None
        if not fn_cell and not fl_cell:
            continue
        raw_fn = str(fn_cell).strip() if fn_cell else str(fl_cell).strip()

        # 策略1: 括号前的部分往往是真正的字段名，如 "FNumber（单据编码）" → "FNumber"
        field_name = re.split(r"[（(（\s]", raw_fn)[0].strip()

        # 策略2: 顿号分隔的多字段名（如 "FYear、FMonth"）合并为下划线连接
        if "、" in field_name:
            parts = [p.strip() for p in field_name.split("、") if p.strip()]
            # 取所有 ASCII 部分拼接
            ascii_parts = [re.findall(r"[A-Za-z][A-Za-z0-9_]*", p) for p in parts]
            flat = [item for sub in ascii_parts for item in sub]
            field_name = "_".join(flat) if flat else field_name.replace("、", "_")

        # 策略3: 若仍含中文/特殊字符，尝试提取连续的英文+数字+下划线段（最长优先）
        if re.search(r"[^\x00-\x7F]", field_name):
            candidates = re.findall(r"[A-Za-z_][A-Za-z0-9_]*", raw_fn)
            if candidates:
                field_name = max(candidates, key=len)
            else:
                # 全是中文：保留中文名，后续通过 Claude API 翻译
                field_name = field_name  # 保持中文，apply_translations 会用 field_label 翻译

        field_name = field_name.strip("_").strip()
        if not field_name or field_name.startswith("#") or field_name in ("-", "—"):
            continue
        if len(field_name) > 100:
            continue

        # 转换为 snake_case（保留全大写缩写如 FXMBH 不变，混合驼峰转下划线）
        field_name = to_snake_case(field_name)

        def get(key):
            c = col.get(key)
            if c is None:
                return ""
            v = ws.cell(row=row_idx, column=c).value
            return str(v).strip() if v is not None else ""

        # 分组列：决定归哪张表
        group_val = get("group")
        if group_val:
            last_group = group_val
        else:
            group_val = last_group  # 合并单元格时分组列为空，继承上一行

        # 根据分组判断表类型（仅在没有独立分区标题行的情况下生效）
        # 分组名含以下关键词 → 主表；其余 → 明细表（detail）
        HEAD_GROUP_KEYWORDS = ("表头", "单头", "基础信息", "基本信息")
        if "明细表#" not in current_table_type and "info" not in current_table_type.lower():
            if group_val == "" or any(k in group_val for k in HEAD_GROUP_KEYWORDS):
                current_table_type = "主表"
            else:
                current_table_type = "明细表"

        field_label = get("field_label")
        raw_type    = get("field_type")
        length      = get("length")
        required_str= get("required")
        enum_values = get("enum_values")
        remark      = get("remark")

        required = required_str in ("是", "Y", "y", "1", "true", "True")
        mysql_type = resolve_type(raw_type, field_label)

        fields.append({
            "table_type":   current_table_type,
            "field_name":   field_name,
            "field_label":  field_label,
            "field_type":   mysql_type,
            "length":       length,
            "required":     required,
            "enum_values":  enum_values,
            "remark":       remark,
        })

    return fields


def translate_labels_to_snake_case(labels: list[str]) -> dict[str, str]:
    """
    调用 Claude API，将中文字段标签批量翻译为 snake_case 英文字段名。
    返回 {label: snake_case_name} 映射。
    若 API 不可用，回退到原始英文缩写处理。
    """
    try:
        import anthropic
    except ImportError:
        print("[WARN] anthropic 未安装，跳过翻译。pip install anthropic", file=sys.stderr)
        return {}

    api_key = (
        os.environ.get("ANTHROPIC_API_KEY")
        or os.environ.get("ANTHROPIC_AUTH_TOKEN")
        or ""
    )
    base_url = os.environ.get("ANTHROPIC_BASE_URL", "https://api.anthropic.com")
    if not api_key:
        print("[WARN] ANTHROPIC_API_KEY 未设置，跳过翻译", file=sys.stderr)
        return {}

    # 去重
    unique_labels = list(dict.fromkeys(labels))

    prompt = (
        "将以下中文数据库字段名称翻译为英文 snake_case 格式。\n"
        "规则：\n"
        "- 简洁准确，反映字段含义\n"
        "- 全部小写，单词间用下划线分隔\n"
        "- 常见缩写可保留：id、no、num、qty、amt、tax、date、type、code、name、status、flag\n"
        "- 直接返回 JSON 对象，格式：{\"中文名\": \"snake_case_name\", ...}\n"
        "- 不要有任何解释，只输出 JSON\n\n"
        "字段列表：\n"
        + json.dumps(unique_labels, ensure_ascii=False)
    )

    try:
        client = anthropic.Anthropic(api_key=api_key, base_url=base_url)
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=4096,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = message.content[0].text.strip()
        # 提取 JSON 块
        m = re.search(r"\{[\s\S]+\}", raw)
        if m:
            return json.loads(m.group())
    except Exception as e:
        print(f"[WARN] 翻译 API 调用失败: {e}", file=sys.stderr)

    return {}


def apply_translations(fields: list, translation: dict) -> list:
    """
    用翻译结果更新每个字段的 field_name。
    若翻译中没有对应 label，保留原来的 field_name（清洗后的英文缩写）。
    同时对重复的翻译名自动加数字后缀。
    """
    seen = {}
    result = []
    for f in fields:
        label = f.get("field_label", "").strip()
        translated = translation.get(label, "").strip()

        if translated:
            # 确保是合法 snake_case
            name = re.sub(r"[^a-z0-9_]", "_", translated.lower()).strip("_")
            name = re.sub(r"_+", "_", name)
        else:
            name = f["field_name"]  # 回退到清洗后的原始缩写

        # 处理重复
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1

        result.append({**f, "field_name": name})
    return result


def parse_excel(excel_path: str, sheet_filter: str = None) -> list:
    """
    入口：读取 Excel，返回所有字段说明列表（field_name 为翻译后的 snake_case）。
    sheet_filter: 若指定，只处理 sheet 名包含该字符串的 sheet。
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_names = find_field_sheets(wb)

    if not sheet_names:
        print(f"[ERROR] 未找到包含'字段说明'的 sheet，已有 sheet: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    if sheet_filter:
        sheet_names = [n for n in sheet_names if sheet_filter in n]
        if not sheet_names:
            print(f"[ERROR] 没有匹配 '{sheet_filter}' 的字段说明 sheet", file=sys.stderr)
            sys.exit(1)
        print(f"[INFO] 使用 sheet: {sheet_names}", file=sys.stderr)

    all_fields = []
    for name in sheet_names:
        fields = parse_field_sheet(wb[name], name)
        all_fields.extend(fields)

    # 批量翻译 field_label → snake_case 英文字段名
    labels = [f.get("field_label", "") for f in all_fields]
    print(f"[INFO] 正在翻译 {len(set(labels))} 个字段名...", file=sys.stderr)
    translation = translate_labels_to_snake_case(labels)
    if translation:
        print(f"[INFO] 翻译完成，覆盖 {len(translation)} 个字段", file=sys.stderr)
        all_fields = apply_translations(all_fields, translation)
    else:
        print("[INFO] 翻译不可用，使用原始英文缩写", file=sys.stderr)

    return all_fields


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parse_excel.py <excel_path> [sheet_filter]", file=sys.stderr)
        sys.exit(1)

    sheet_filter = sys.argv[2] if len(sys.argv) >= 3 else None
    result = parse_excel(sys.argv[1], sheet_filter)
    print(json.dumps(result, ensure_ascii=False, indent=2))