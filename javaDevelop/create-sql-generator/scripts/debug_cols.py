import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from parse_excel import find_field_sheets, find_header_row, parse_col_indices

path = sys.argv[1] if len(sys.argv) > 1 else r"E:\backstage\00Cluade\01skillUse\construct-star-server\.claude\skills\create-sql-generator\source\固定资产入库.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)

for name in find_field_sheets(wb):
    ws = wb[name]
    hr = find_header_row(ws)
    print(f"Sheet: {name}, header_row={hr}")
    if hr:
        cols = parse_col_indices(ws, hr)
        print(f"  col indices: {cols}")
        # print header row raw values
        row_vals = [(c, ws.cell(row=hr, column=c).value) for c in range(1, ws.max_column+1) if ws.cell(row=hr, column=c).value]
        print(f"  header cells: {row_vals}")
        # print first 5 data rows
        print("  first 5 data rows (field_name col, field_label col):")
        fn_col = cols.get("field_name")
        fl_col = cols.get("field_label")
        grp_col = cols.get("group")
        for ri in range(hr+1, min(hr+8, ws.max_row+1)):
            fn = ws.cell(row=ri, column=fn_col).value if fn_col else None
            fl = ws.cell(row=ri, column=fl_col).value if fl_col else None
            grp = ws.cell(row=ri, column=grp_col).value if grp_col else None
            print(f"    row {ri}: group={grp!r}  field_name={fn!r}  field_label={fl!r}")
