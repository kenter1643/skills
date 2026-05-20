# -*- coding: utf-8 -*-
"""
Excel 解析器 - 独立可执行脚本
解析 Excel 的 sheet 页：xxx-字段说明
输出 JSON 到 output/ 目录

用法:
    python parse_excel.py <excel_file> [--output <output_dir>]
"""

import argparse
import json
import os
import re
import sys

import openpyxl


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------

FIELD_NAME_ALIASES = {
    '所属组织名称': '所属组织',
    '账套': '账套名称',
    '所属账套': '账套名称',
    '账套名': '账套名称',
    '制单日期': '制单时间',
}

FIELD_PAIR_MAPPINGS = {
    '所属组织': '所属组织编号',
    '所属组织编号': '所属组织',
    '账套名称': '账套编号',
    '账套编号': '账套名称',
}


def _norm(s):
    return re.sub(r'[（）]', lambda m: '(' if m.group() == '（' else ')', str(s).strip())


def canonicalize_field_name(name):
    normalized = _norm(name)
    return FIELD_NAME_ALIASES.get(normalized, normalized)


def expand_related_field_names(names):
    """按成对规则补齐关联字段名，并保持原有顺序"""
    result = []
    seen = set()
    for name in names:
        cn = canonicalize_field_name(name)
        if cn not in seen:
            seen.add(cn)
            result.append(cn)
        related = FIELD_PAIR_MAPPINGS.get(cn)
        if related and related not in seen:
            seen.add(related)
            result.append(related)
    return result


# ---------------------------------------------------------------------------
# 表单展现映射 & 查询方式解析
# ---------------------------------------------------------------------------

DISPLAY_TYPE_MAP = [
    # 注意：更长/更具体的匹配项必须放在前面，避免被子串误匹配
    ('多行文本框', 'textarea'),
    ('多行文本', 'textarea'),
    ('下拉', 'select'),
    ('复选框', 'checkbox'),
    ('单选框', 'radio'),
    ('附件', 'file'),
    ('按钮', 'button'),
    ('抽屉', 'drawer'),
    ('弹窗', 'dialog'),
    ('导入', 'import'),
    ('日期', 'date'),
    ('数字', 'number'),
    ('文本', 'text'),
]


def map_display_type(raw_value):
    if not raw_value:
        return ''
    raw_str = str(raw_value).strip()
    for keyword, mapped in DISPLAY_TYPE_MAP:
        if keyword in raw_str:
            return mapped
    return raw_str


def is_hidden_row(raw_display):
    if not raw_display:
        return False
    return '隐藏' in str(raw_display).strip()


def parse_query_method(raw_value):
    if not raw_value:
        return False, False
    raw_str = str(raw_value).strip()
    return '模糊' in raw_str, '高级' in raw_str


# ---------------------------------------------------------------------------
# 字段说明 sheet 解析
# ---------------------------------------------------------------------------

def _normalize_name(name):
    return canonicalize_field_name(name)


def _find_col_indices(ws):
    """动态扫描字段说明 sheet 的表头行，返回关键列的索引（0-based）"""
    keyword_map = {
        '分组信息': 'group',
        '字段名称': 'field_name',
        '字段名': 'field_name',
        '字段类型': 'field_type',
        '表单展现': 'display_type',
        '数据生成方式': 'data_generation',
        '数据检查要求': 'data_check',
        '数据检查要求（校验规则）': 'data_check',
        '字段说明': 'field_requirement',
        '是否列表字段': 'is_list_field',
        '列表顺序': 'list_order',
        '查询方式': 'query_method',
        '是否可编辑': 'is_editable',
        '是否必填': 'is_required',
        '是否枚举': 'is_enum',
        '枚举内容': 'enum_values',
        '枚举值': 'enum_values',
    }
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        indices = {}
        for col_idx, cell_val in enumerate(row):
            if not cell_val:
                continue
            cell_str = str(cell_val).strip()
            for keyword, key in keyword_map.items():
                if keyword in cell_str:
                    indices[key] = col_idx
        if 'field_name' in indices:
            return indices
    return None


def parse_field_description_sheet(ws):
    """解析字段说明 sheet，按规则提取所有字段信息"""
    fields = {}
    has_attachment = False
    field_order = []
    fuzzy_search = []
    advanced_search = []
    list_fields_with_order = []

    col_indices = _find_col_indices(ws)
    if col_indices is None:
        return fields, has_attachment, field_order, fuzzy_search, advanced_search, []

    fn_col = col_indices.get('field_name')
    grp_col = col_indices.get('group')
    ft_col = col_indices.get('field_type')
    dt_col = col_indices.get('display_type')
    dg_col = col_indices.get('data_generation')
    dc_col = col_indices.get('data_check')
    fr_col = col_indices.get('field_requirement')
    lf_col = col_indices.get('is_list_field')
    lo_col = col_indices.get('list_order')
    qm_col = col_indices.get('query_method')
    ed_col = col_indices.get('is_editable')
    req_col = col_indices.get('is_required')
    enum_col = col_indices.get('is_enum')
    eval_col = col_indices.get('enum_values')

    def _parse_bool(val):
        return str(val).strip() == '是' if val else False

    def _parse_int(val):
        try:
            return int(val)
        except (ValueError, TypeError):
            return 0

    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        if fn_col is not None and fn_col < len(row) and row[fn_col] and '字段' in str(row[fn_col]):
            header_row = row_idx
            break
    data_start = (header_row + 1) if header_row else 2

    current_group = ''
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if fn_col is None or fn_col >= len(row):
            continue

        raw_name = row[fn_col]
        if not raw_name:
            continue

        # 先提取分组信息（规则1），在过滤隐藏行之前做 fill-down
        row_group = ''
        if grp_col is not None and grp_col < len(row):
            row_group = str(row[grp_col] or '')
        if row_group:
            current_group = row_group
        group = current_group

        field_type = ''
        if ft_col is not None and ft_col < len(row):
            field_type = str(row[ft_col] or '')

        raw_display = ''
        if dt_col is not None and dt_col < len(row):
            raw_display = str(row[dt_col] or '')

        # 规则2：表单展现包含"隐藏"的行过滤掉
        if is_hidden_row(raw_display):
            continue

        field_name = _normalize_name(raw_name)

        # 规则3：映射 display_type
        display_type = map_display_type(raw_display)

        data_generation = ''
        if dg_col is not None and dg_col < len(row):
            data_generation = str(row[dg_col] or '')

        data_check = ''
        if dc_col is not None and dc_col < len(row):
            data_check = str(row[dc_col] or '')

        field_requirement = ''
        if fr_col is not None and fr_col < len(row):
            field_requirement = str(row[fr_col] or '')

        # 规则3：是否列表字段 + 列表顺序
        is_list_field = False
        list_order = 0
        if lf_col is not None and lf_col < len(row):
            is_list_field = _parse_bool(row[lf_col])
            if is_list_field and lo_col is not None and lo_col < len(row):
                list_order = _parse_int(row[lo_col])

        # 规则4：查询方式列
        is_fuzzy_search = False
        is_advanced_search = False
        if qm_col is not None and qm_col < len(row):
            is_fuzzy_search, is_advanced_search = parse_query_method(row[qm_col])

        # 规则5：是否可编辑
        is_editable = False
        if ed_col is not None and ed_col < len(row):
            is_editable = _parse_bool(row[ed_col])

        # 规则6：是否必填
        is_required = False
        if req_col is not None and req_col < len(row):
            is_required = _parse_bool(row[req_col])

        # 规则7：是否枚举
        is_enum = False
        if enum_col is not None and enum_col < len(row):
            is_enum = _parse_bool(row[enum_col])

        enum_values = ''
        if eval_col is not None and eval_col < len(row):
            enum_values = str(row[eval_col] or '')

        if '附件' in str(raw_name) or '文件' in str(raw_name):
            has_attachment = True

        fields[field_name] = {
            'group': group,
            'field_type': field_type,
            'display_type': display_type,
            'data_generation': data_generation,
            'data_check': data_check,
            'field_requirement': field_requirement,
            'is_list_field': is_list_field,
            'list_order': list_order,
            'is_fuzzy_search': is_fuzzy_search,
            'is_advanced_search': is_advanced_search,
            'is_editable': is_editable,
            'is_required': is_required,
            'is_enum': is_enum,
            'enum_values': enum_values,
        }
        if field_name not in field_order:
            field_order.append(field_name)

        if is_fuzzy_search:
            fuzzy_search.append(field_name)
            if field_name not in advanced_search:
                advanced_search.append(field_name)
        elif is_advanced_search:
            if field_name not in advanced_search:
                advanced_search.append(field_name)

        if is_list_field:
            list_fields_with_order.append((field_name, list_order))

    # 同步 field_info 中的 is_fuzzy_search / is_advanced_search
    # 与最终输出的 fuzzy_search / advanced_search 列表保持一致
    fuzzy_set = set(fuzzy_search)
    advanced_set = set(advanced_search)
    for field_name, info in fields.items():
        info['is_fuzzy_search'] = field_name in fuzzy_set
        info['is_advanced_search'] = field_name in advanced_set

    # 规则3：按列表顺序排序，输出 list_order_fields
    list_fields_with_order.sort(key=lambda x: x[1])
    list_order_fields = expand_related_field_names([name for name, _ in list_fields_with_order])
    # 过滤掉被隐藏的字段（不在 fields 字典中的字段）
    list_order_fields = [f for f in list_order_fields if f in fields]

    return fields, has_attachment, field_order, fuzzy_search, advanced_search, list_order_fields


# ---------------------------------------------------------------------------
# input_style 生成（基于字段说明的分组信息）
# ---------------------------------------------------------------------------

def _parse_drawer_fields(text):
    """从字段说明文本中解析抽屉的模糊/高级/列表字段"""
    result = {'fuzzy_search': [], 'advanced_search': [], 'list_fields': []}
    if not text:
        return result

    # 提取中文顿号分隔的字段名
    def extract_fields(segment):
        if not segment:
            return []
        items = []
        for part in re.split(r'[；;，,。]', segment):
            part = part.strip()
            if not part:
                continue
            for item in re.split(r'[、]', part):
                item = item.strip()
                if item and len(item) <= 30 and item not in items:
                    items.append(item)
        return items

    # 模糊查询：XXX、XXX
    m = re.search(r'模糊查询[：:](.+?)(?:[；;]|$)', text)
    if m:
        result['fuzzy_search'] = extract_fields(m.group(1))
        for f in result['fuzzy_search']:
            if f not in result['advanced_search']:
                result['advanced_search'].append(f)

    # 高级查询：XXX、XXX
    m = re.search(r'高级查询[：:](.+?)(?:[；;]|$)', text)
    if m:
        for f in extract_fields(m.group(1)):
            if f not in result['advanced_search']:
                result['advanced_search'].append(f)

    # 列表字段 / 抽屉内字段
    m = re.search(r'(?:抽屉内字段|列表字段)[：:](.+?)(?:[；;]|$)', text)
    if m:
        result['list_fields'] = extract_fields(m.group(1))

    return result


# 分组内的字段排序规则（按此顺序排列，未列出的字段按原始顺序排在最后）
GROUP_FIELD_ORDER = {
    '基础信息': ['承办项目/部门', '制单时间', '制单人'],
}


def _sort_group_fields(title, fields):
    """按 GROUP_FIELD_ORDER 规则对分组内字段排序"""
    order = GROUP_FIELD_ORDER.get(title, [])
    if not order:
        return fields
    priority = {name: i for i, name in enumerate(order)}
    return sorted(fields, key=lambda f: priority.get(f, len(order)))


def build_input_style(fields, field_order):
    """根据字段说明 sheet 的分组信息构建 input_style"""
    # 填充合并单元格导致的空分组（向下填充）
    filled_groups = {}
    current_group = ''
    for field_name in field_order:
        info = fields.get(field_name)
        if info is None:
            continue
        group = info.get('group', '')
        if group:
            current_group = group
        filled_groups[field_name] = current_group

    # 按分组归类
    group_map = {}
    for field_name in field_order:
        group = filled_groups.get(field_name, '')
        if group not in group_map:
            group_map[group] = []
        group_map[group].append(field_name)

    main_groups = []
    details = []
    detail_title = '明细'
    drawers = []
    detail_seen = set()

    for title in field_order:
        # 用第一个出现的字段的分组来确定group顺序，但我们需要按group去重
        pass

    # 按首次出现的顺序输出分组
    seen_groups = set()
    ordered_groups = []
    for field_name in field_order:
        group = filled_groups.get(field_name, '')
        if group not in seen_groups:
            seen_groups.add(group)
            ordered_groups.append(group)

    for title in ordered_groups:
        group_fields = group_map.get(title, [])
        if not group_fields:
            continue
        if '明细' in title:
            for f in group_fields:
                if f not in detail_seen:
                    detail_seen.add(f)
                    details.append(f)
            detail_title = title
        elif title:
            main_groups.append({'title': title, 'fields': _sort_group_fields(title, group_fields)})
        else:
            # 无分组字段归入默认组
            main_groups.append({'title': '', 'fields': group_fields})

    # 从 display_type 为 drawer 的字段中提取抽屉信息
    for field_name in field_order:
        info = fields.get(field_name)
        if info is None or info.get('display_type') != 'drawer':
            continue
        drawer = _parse_drawer_fields(info.get('field_requirement', ''))
        drawer['title'] = field_name
        drawers.append(drawer)

    return {
        'main': main_groups,
        'details': {'title': detail_title, 'fields': details},
        'drawers': drawers,
    }


# ---------------------------------------------------------------------------
# 主解析函数
# ---------------------------------------------------------------------------

def parse_excel(excel_path, business_name=None):
    if business_name is None:
        business_name = os.path.splitext(os.path.basename(excel_path))[0]

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # 查找字段说明 sheet
    field_sheet_name = None
    for sheet_name in wb.sheetnames:
        if '字段说明' in sheet_name:
            field_sheet_name = sheet_name
            break

    if not field_sheet_name:
        raise ValueError(f"未找到字段说明 sheet，请确认 Excel 中包含 '字段说明' 的 sheet")

    # 解析字段说明 sheet
    field_info, has_attachment, field_order, fuzzy_search, advanced_search, list_order_fields = \
        parse_field_description_sheet(wb[field_sheet_name])

    # 根据分组信息构建 input_style
    input_style = build_input_style(field_info, field_order)

    # 计算 main_fields 和 detail_fields
    main_fields = []
    detail_fields = []
    for g in input_style['main']:
        main_fields.extend(g['fields'])
    detail_fields = list(input_style['details']['fields'])

    wb.close()

    return {
        'business_name': business_name,
        'fuzzy_search': fuzzy_search,
        'advanced_search': advanced_search,
        'list_order_fields': list_order_fields,
        'field_info': field_info,
        'main_fields': main_fields,
        'detail_fields': detail_fields,
        'input_style': input_style,
        'has_attachment': has_attachment,
        'field_description_sheet': field_sheet_name,
    }


# ---------------------------------------------------------------------------
# 命令行入口
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Excel 解析器 - 将 Excel 需求文档解析为 JSON',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
    python parse_excel.py "固定资产销售.xlsx"
    python parse_excel.py "固定资产销售.xlsx" --output ./my_output
    python parse_excel.py "固定资产销售.xlsx" --business-name "固定资产销售" --compact
        '''
    )
    parser.add_argument('excel_file', help='Excel 文件路径')
    parser.add_argument('--output', '-o', default='output', help='输出目录（默认：output）')
    parser.add_argument('--business-name', '-b', default=None, help='业务名称（默认从文件名提取）')
    parser.add_argument('--compact', '-c', action='store_true', help='紧凑 JSON 输出')

    args = parser.parse_args()

    if not os.path.exists(args.excel_file):
        print(f"[错误] Excel 文件不存在: {args.excel_file}", file=sys.stderr)
        sys.exit(1)

    business_name = args.business_name or os.path.splitext(os.path.basename(args.excel_file))[0]

    print(f"[解析] Excel 文件: {args.excel_file}")
    print(f"[解析] 业务名称: {business_name}")

    data = parse_excel(args.excel_file, business_name)

    # 打印摘要
    print(f"[结果] 模糊查询字段: {len(data['fuzzy_search'])} 个")
    print(f"[结果] 高级查询字段: {len(data['advanced_search'])} 个")
    print(f"[结果] 列表字段: {len(data['list_order_fields'])} 个")
    print(f"[结果] 字段说明: {len(data['field_info'])} 个")
    print(f"[结果] 主表字段: {len(data.get('main_fields', []))} 个")
    print(f"[结果] 明细表字段: {len(data.get('detail_fields', []))} 个")
    main_fields_count = sum(len(g['fields']) for g in data['input_style']['main'])
    print(f"[结果] 主信息区: {main_fields_count} 个字段（{len(data['input_style']['main'])} 个分组）")
    print(f"[结果] 明细区: {len(data['input_style']['details']['fields'])} 个字段")
    print(f"[结果] 抽屉区域: {len(data['input_style'].get('drawers', []))} 个")

    # 创建输出目录
    os.makedirs(args.output, exist_ok=True)

    # 输出 JSON
    output_filename = f"{business_name}.json"
    output_path = os.path.join(args.output, output_filename)

    indent = 2 if not args.compact else None
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=indent, ensure_ascii=False)

    print(f"[输出] JSON 已保存至: {os.path.abspath(output_path)}")

    # 输出字段摘要文本文件
    summary_path = os.path.join(args.output, f"{business_name}_summary.txt")
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write(f"业务名称: {business_name}\n")
        f.write(f"来源文件: {args.excel_file}\n")
        f.write(f"字段说明 sheet: {data['field_description_sheet']}\n")
        f.write(f"\n--- 模糊查询字段 ({len(data['fuzzy_search'])} 个) ---\n")
        for name in data['fuzzy_search']:
            f.write(f"  {name}\n")
        f.write(f"\n--- 高级查询字段 ({len(data['advanced_search'])} 个) ---\n")
        for name in data['advanced_search']:
            f.write(f"  {name}\n")
        f.write(f"\n--- 列表字段 ({len(data['list_order_fields'])} 个) ---\n")
        for name in data['list_order_fields']:
            f.write(f"  {name}\n")
        f.write(f"\n--- 主表字段 ({len(data.get('main_fields', []))} 个) ---\n")
        for name in data.get('main_fields', []):
            f.write(f"  {name}\n")
        f.write(f"\n--- 明细表字段 ({len(data.get('detail_fields', []))} 个) ---\n")
        for name in data.get('detail_fields', []):
            f.write(f"  {name}\n")
        f.write(f"\n--- 字段说明详情 ({len(data['field_info'])} 个) ---\n")
        for name, info in data['field_info'].items():
            parts = []
            if info.get('group'):
                parts.append(f"分组={info['group']}")
            parts.append(f"必填={info['is_required']}")
            parts.append(f"枚举={info['is_enum']}")
            if info.get('enum_values'):
                parts.append(f"枚举值={info['enum_values']}")
            if info.get('field_type'):
                parts.append(f"类型={info['field_type']}")
            if info.get('display_type'):
                parts.append(f"展现={info['display_type']}")
            if info.get('is_list_field'):
                parts.append(f"列表字段(顺{info.get('list_order',0)})")
            if info.get('is_fuzzy_search'):
                parts.append(f"模糊查询")
            if info.get('is_advanced_search'):
                parts.append(f"高级查询")
            if info.get('is_editable'):
                parts.append(f"可编辑")
            f.write(f"  {name}: {', '.join(parts)}\n")
            if info.get('data_generation'):
                f.write(f"    数据生成: {info['data_generation'][:80]}\n")
            if info.get('data_check'):
                f.write(f"    数据检查: {info['data_check'][:80]}\n")
            if info.get('field_requirement'):
                f.write(f"    字段要求: {info['field_requirement'][:80]}\n")
        f.write(f"\n--- 主信息区（{len(data['input_style']['main'])} 个分组）---\n")
        for group in data['input_style']['main']:
            title = group['title'] or '(未分组)'
            f.write(f"  [{title}]\n")
            for name in group['fields']:
                f.write(f"    {name}\n")
        f.write(f"\n--- {data['input_style']['details']['title']} ({len(data['input_style']['details']['fields'])} 个) ---\n")
        for name in data['input_style']['details']['fields']:
            f.write(f"  {name}\n")
        for dr in data['input_style'].get('drawers', []):
            f.write(f"\n--- 抽屉: {dr['title']} ---\n")
            f.write(f"  模糊查询: {dr['fuzzy_search']}\n")
            f.write(f"  高级查询: {dr['advanced_search']}\n")
            f.write(f"  列表字段: {dr['list_fields']}\n")

    print(f"[输出] 摘要已保存至: {os.path.abspath(summary_path)}")


if __name__ == '__main__':
    main()
